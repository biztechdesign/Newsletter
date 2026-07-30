"""
import_content.py
=================
Auto-generates content.json from two Google Sheets:
  1. Newsletter Content sheet  — text, awards, openings, blogs
  2. Work Anniversary sheet    — employee anniversary data

Both sheets must be shared as "Anyone with the link can view".

Usage:
    python import_content.py          → content.json + browser-preview HTML
    python import_content.py --all    → also runs steps 2 and 4 below

This stops at the browser preview so the content can be checked before the
section images are cut. The full monthly flow is:

    1. python import_content.py         → content.json + preview HTML
       ...review the preview, fix the sheet/images and re-run until approved
    2. python screenshot_sections.py    → PNGs in "<Month>-<Year>/Section wise images"
    3. upload those PNGs                → see UPLOAD_ME.txt in that same folder
    4. python generate_simple_email.py  → final Gmail/Outlook-safe email HTML

This month's source photos go in "<Month>-<Year>/Row Data" (awards/,
new_joinee/, events/event1/, …) — see ROW_DATA_FOLDERS for the full list.
Those folder names are the only wiring needed, so the sheet's Drive Link
column can be left empty for images; anything found locally wins over a link.

Update the SHEET IDs at the top when working on a new month.
"""

import io
import json
import os
import re
import subprocess
import sys
import urllib.error
import urllib.request
from pathlib import Path
from urllib.parse import quote

import openpyxl

# ── Configure these each month ────────────────────────────────────────────
CONTENT_SHEET_ID = "1vZhE7L6O7NNAHYrsGOm8nlZjjwQtMzbu"
# Anniversary file + tab are now read from the 'anniversary' row in the
# content sheet (Drive Link column = xlsx file, Value column = tab name).
# ANNIV_SHEET_ID below is only used as a fallback if not specified in sheet.
ANNIV_SHEET_ID   = "19m9QWBhOqLh9pZT-u4eoaPkN9XS3hu69"
# ──────────────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
OUTPUT   = BASE_DIR / "content.json"

# This month's local source-image folder, e.g. "July-2026/Row Data". Set by
# main() once the sheet's month/year are known. Images found here take
# priority over anything linked in the sheet — see ROW_DATA_FOLDERS.
ROW_DATA_DIR: Path | None = None

# Cached by github_token(): "" means "looked and found nothing", None means
# "not looked yet".
_GITHUB_TOKEN: str | None = None

# Footer is static — update here if it ever changes
FOOTER = {
    "address": "C/801 Dev Aurum Commercial, Anandnagar Cross Road Prahalad Nagar, Satellite, Ahmedabad, Gujarat 380015",
    "email":   "career@biztechcs.com",
    "phone":   "+91 93276 55844",
    "social": {
        "linkedin":  "https://www.linkedin.com/company/biztech/",
        "facebook":  "https://www.facebook.com/biztech/",
        "youtube":   "https://www.youtube.com/@Biztechcs",
        "instagram": "https://www.instagram.com/biztechcs/",
        "twitter":   "https://twitter.com/biztechcs",
        "dribbble":  "https://dribbble.com/biztechcs",
    },
}


# ── Helpers ───────────────────────────────────────────────────────────────

def download_xlsx(sheet_id: str) -> openpyxl.Workbook:
    """Download a Google Sheet as xlsx (follows redirects automatically)."""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req) as resp:
        return openpyxl.load_workbook(io.BytesIO(resp.read()))


def load_xlsx(sheet_id: str, local_fallback: str) -> openpyxl.Workbook:
    """
    Try to download from Google Sheets. If the sheet is private (401),
    fall back to reading a local xlsx file.
    """
    try:
        wb = download_xlsx(sheet_id)
        print("  Downloaded from Google Sheets.")
        return wb
    except urllib.error.HTTPError as e:
        if e.code == 401:
            local = BASE_DIR / local_fallback
            if local.exists():
                print(f"  Sheet is private — reading local file: {local_fallback}")
                return openpyxl.load_workbook(local)
            else:
                raise FileNotFoundError(
                    f"\n  The sheet is not publicly shared and no local file found.\n"
                    f"  Fix option 1: Share the sheet → 'Anyone with the link can view'\n"
                    f"  Fix option 2: Download it as .xlsx and save as: {local}\n"
                ) from e
        raise


def parse_years(tenure_str) -> int:
    """Extract numeric years from '14 years' or '1 year'."""
    m = re.search(r"(\d+)", str(tenure_str or ""))
    return int(m.group(1)) if m else 0


# Written in a cell to mean "nothing this month". Treated exactly like an
# empty cell, so the section drops out of the newsletter instead of rendering
# the word itself as if it were content.
NOT_APPLICABLE = {"na", "n/a", "n.a.", "n.a", "-", "--", "none", "nil",
                  "not applicable", "no", "x"}

# Which section each sheet field belongs to. Writing NA against any of these
# drops that whole section from the newsletter, even if its Row Data folder
# still has images in it — "NA" is an explicit instruction, not just an
# absence, so it has to beat leftover files.
FIELD_SECTION = {
    "ceo_message":            "ceo",
    "campaign_banner":        "campaign",
    "new_customer":           "new_customers",
    "delivery":               "delivery",
    "delivery_description":   "delivery",
    "excellence":             "excellence",
    "hr_wellness_banner":     "hr",
    "hr_event": "hr", "hr_event1": "hr", "hr_event2": "hr", "hr_event3": "hr",
    "award":                  "rewards",
    "employee_certification": "certifications",
    "anniversary":            "anniversaries",
    "new_joinee":             "new_joinees",
    "opening":                "openings",
    "new_openings_sheet":     "openings",
    "blog1": "marketing", "blog2": "marketing",
    "marketing_highlights":   "marketing",
    "new_announcement":       "announcements",
    "expo":                   "expo",
    "employee_training":      "training",
    "employee_workshop":      "workshop",
}


def is_na(value) -> bool:
    """Whether a cell says 'nothing this month' rather than holding content."""
    return isinstance(value, str) and value.strip().lower() in NOT_APPLICABLE


def blank_if_na(value):
    """
    None for a cell that says 'NA', otherwise the value untouched.

    Deliberately does not stringify: some cells hold real dates (the
    anniversary row's month) or numbers, and the callers depend on those
    types surviving.
    """
    return None if is_na(value) else value


def cell_str(value) -> str:
    """
    Safe string conversion of a cell value, with 'NA' and friends normalised
    to empty so they read as 'no content this month' everywhere downstream.
    """
    if value is None or is_na(value):
        return ""
    return str(value).strip()


IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg"}


def drive_file_id(drive_url: str) -> str:
    """Extract the file ID from any Google Drive URL."""
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", drive_url or "")
    return m.group(1) if m else ""


# ── GitHub helpers ────────────────────────────────────────────────────────

def parse_github_url(url: str) -> dict | None:
    """
    Parse a github.com tree/blob URL.
    Returns {owner, repo, branch, path, is_folder} or None.
    """
    m = re.match(r"https://github\.com/([^/]+)/([^/]+)/(tree|blob)/([^/]+)/(.+)", url or "")
    if m:
        return {
            "owner":     m.group(1),
            "repo":      m.group(2),
            "is_folder": m.group(3) == "tree",
            "branch":    m.group(4),
            "path":      m.group(5).rstrip("/"),
        }
    return None


def github_token() -> str:
    """
    A GitHub token, if one can be found, cached after the first lookup.

    Unauthenticated API calls are capped at 60/hour and each run makes roughly
    a dozen folder listings, so a few runs in a row start returning 403 and
    whole sections silently vanish from the newsletter. An authenticated call
    gets 5000/hour. Checked in order: GITHUB_TOKEN, GH_TOKEN, then the GitHub
    CLI if it's installed and logged in.
    """
    global _GITHUB_TOKEN
    if _GITHUB_TOKEN is not None:
        return _GITHUB_TOKEN

    _GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN") or os.environ.get("GH_TOKEN") or ""
    if not _GITHUB_TOKEN:
        try:
            result = subprocess.run(
                ["gh", "auth", "token"],
                capture_output=True, text=True, timeout=10,
            )
            if result.returncode == 0:
                _GITHUB_TOKEN = result.stdout.strip()
        except (OSError, subprocess.SubprocessError):
            pass  # gh not installed or not logged in — carry on unauthenticated
    return _GITHUB_TOKEN


def list_github_folder(owner: str, repo: str, branch: str, path: str) -> list[dict]:
    """
    List image files in a GitHub repo folder via the API.
    Returns [{name, stem, url}] sorted by name, where url is the GitHub Pages URL.

    Only used as a fallback — this month's Row Data folder is consulted first,
    by convention, in resolve_all_images().
    """
    api = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}?ref={branch}"
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/vnd.github.v3+json",
    }
    token = github_token()
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(api, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            files = json.loads(resp.read())
            result = []
            for f in files:
                if f["type"] == "file" and Path(f["name"]).suffix.lower() in IMAGE_EXTS:
                    pages_url = f"https://{owner}.github.io/{repo}/{path}/{f['name']}"
                    result.append({"name": f["name"], "stem": Path(f["name"]).stem, "url": pages_url})
            return sorted(result, key=lambda x: x["name"])
    except Exception as e:
        print(f"  WARNING: Could not list GitHub folder '{path}': {e}")
        return []


def resolve_single_url(raw: str) -> str:
    """
    Resolve any raw URL to a direct image URL:
    - GitHub folder (tree) → first image in that folder (GitHub Pages URL)
    - GitHub file  (blob)  → GitHub Pages URL
    - Drive share link     → uc?export=view URL
    - Any other http URL   → as-is
    """
    if not raw or "YOUR_FILE_ID" in raw:
        return ""
    info = parse_github_url(raw)
    if info:
        if info["is_folder"]:
            files = list_github_folder(info["owner"], info["repo"], info["branch"], info["path"])
            return files[0]["url"] if files else ""
        else:
            return f"https://{info['owner']}.github.io/{info['repo']}/{info['path']}"
    if "drive.google.com" in raw:
        fid = drive_file_id(raw)
        return f"https://drive.google.com/uc?export=view&id={fid}" if fid else ""
    if raw.startswith("http"):
        return raw
    return ""


def resolve_folder_urls(raw: str) -> list[str]:
    """
    Resolve a folder URL to a list of all image URLs in that folder.
    For non-folder URLs, wraps resolve_single_url() in a list.
    """
    if not raw:
        return []
    info = parse_github_url(raw)
    if info and info["is_folder"]:
        files = list_github_folder(info["owner"], info["repo"], info["branch"], info["path"])
        return [f["url"] for f in files]
    single = resolve_single_url(raw)
    return [single] if single else []


def split_name_parts(stem: str) -> list[str]:
    """
    Split a filename stem into its fields, e.g. 'Name--Designation' or
    '01--Award Title--Name--Designation'.

    A plain hyphen padded with spaces ('Name - Designation') is accepted too,
    since that's what people type naturally. The surrounding spaces are
    required, so hyphenated names like 'Anne-Marie' are never split apart.
    """
    if "--" in stem:
        return [p.strip() for p in stem.split("--")]
    if " - " in stem:
        return [p.strip() for p in stem.split(" - ")]
    return [stem.strip()]


# Where each section's images live under "<Month>-<Year>/Row Data". This is
# the whole contract: drop files in these folders and the sheet needs no image
# links at all. HR events are events/event1, events/event2, … by row order.
ROW_DATA_FOLDERS = {
    "ceo_photo":          "ceo",
    "campaign_banner":    "campaign_banner",
    "hr_wellness_banner": "hr_wellness",
    "delivery":           "delivery",
    "excellence":         "excellence",
    "awards":             "awards",
    "new_joinee":         "new_joinee",
    "new_customer":       "new_customer",
    "certification":      "certification",
    "announcement":       "announcement",
    "blogs":              "marketing/blogs",
    "expo":               "expo",
    "training":           "training",
    "workshop":           "workshop",
}

# Photo-gallery sections that are just a title plus a folder of pictures.
# (data key, sheet field, Row Data folder, label)
GALLERY_SECTIONS = [
    ("expo",     "expo",                "expo",     "Expo / Exhibition"),
    ("training", "employee_training",   "training", "Employee Training"),
    ("workshop", "employee_workshop",   "workshop", "Employee Workshop"),
]


def local_ref(p: Path) -> str:
    """
    How the rendered HTML should reference a local image.

    A path relative to the folder the HTML is written into — NOT an absolute
    file:// URI. The preview is served over http://127.0.0.1 by watch.py, and
    browsers refuse to load file:// subresources from an http:// page, so
    absolute URIs render as broken images there while working fine when opened
    from disk. A relative path works both ways. Spaces and the like are
    percent-encoded, since 'Row Data' and most photo names contain them.
    """
    resolved = p.resolve()
    if not resolved.is_relative_to(BASE_DIR):
        return resolved.as_uri()   # outside the project; nothing better available
    # The HTML lands in "<Month>-<Year>/HTML" next to this month's Row Data, so
    # the path is measured from there rather than assuming a fixed "../" depth.
    html_dir = (ROW_DATA_DIR.parent if ROW_DATA_DIR else BASE_DIR) / "HTML"
    return quote(os.path.relpath(resolved, html_dir).replace("\\", "/"))


def browser_renderable(p: Path) -> bool:
    """
    Whether a browser can actually decode this file, judged by its leading
    bytes rather than its extension.

    Phones shoot HEIC, and renaming one to .png does not convert it — the file
    then passes every extension check, uploads fine, and shows up as a broken
    image in the finished newsletter with nothing to explain why. Cheap to
    catch here, painful to catch by eye later.
    """
    if p.suffix.lower() == ".svg":
        return True
    try:
        head = p.read_bytes()[:16]
    except OSError:
        return False
    if head[:8] == b"\x89PNG\r\n\x1a\n":                     # PNG
        return True
    if head[:3] == b"\xff\xd8\xff":                          # JPEG
        return True
    if head[:6] in (b"GIF87a", b"GIF89a"):                   # GIF
        return True
    if head[:4] == b"RIFF" and head[8:12] == b"WEBP":        # WebP
        return True
    if head[4:8] == b"ftyp" and head[8:12] in (b"avif", b"avis"):
        return True
    return False


def files_in(folder: Path) -> list[dict]:
    """[{name, stem, url}] for images directly inside *folder*, sorted by name."""
    if not folder.is_dir():
        return []
    found = []
    for p in sorted(folder.iterdir(), key=lambda x: x.name):
        if not p.is_file() or p.suffix.lower() not in IMAGE_EXTS:
            continue
        if not browser_renderable(p):
            print(f"  WARNING: '{p.name}' is not a real {p.suffix.lstrip('.').upper()} "
                  f"(looks like HEIC or similar) - skipping; convert it to PNG/JPEG.")
            continue
        found.append({"name": p.name, "stem": p.stem, "url": local_ref(p)})
    return found


def files_from_local_path(raw: str) -> list[dict]:
    """
    Images from a folder named directly in the sheet's Drive Link cell, e.g.
    'Row Data/ceo'. Anything starting with http is a URL, not a path, and is
    left for the GitHub/Drive resolvers.

    Resolved against the month folder first, then the project root, so both
    'Row Data/ceo' and 'July-2026/Row Data/ceo' work, as does an absolute path.
    """
    if not raw or raw.startswith("http"):
        return []
    candidate = Path(raw.replace("\\", "/"))
    if candidate.is_absolute():
        return files_in(candidate)
    roots = []
    if ROW_DATA_DIR is not None:
        roots.append(ROW_DATA_DIR.parent)   # the "<Month>-<Year>" folder
    roots.append(BASE_DIR)
    for root in roots:
        found = files_in(root / candidate)
        if found:
            return found
    return []


def local_files(subfolder: str) -> list[dict]:
    """
    [{name, stem, url}] for the images in a Row Data subfolder, sorted by name
    so numeric filename prefixes control display order. Empty if this month has
    no local folder, or that section's folder is missing or has no images.
    """
    if ROW_DATA_DIR is None:
        return []
    return files_in(ROW_DATA_DIR / subfolder)


def github_folder_files(folder_url: str) -> list[dict]:
    """[{name, stem, url}] for a GitHub folder URL, or [] if it isn't one."""
    info = parse_github_url(folder_url)
    if not info or not info["is_folder"]:
        return []
    return list_github_folder(info["owner"], info["repo"], info["branch"], info["path"])


def parse_awards(files: list[dict]) -> list[dict]:
    """
    Parse filenames as 'Title--Name--Designation.ext', or with an explicit
    display-order prefix, '01--Title--Name--Designation.ext'.
    Return awards list with recipients and image_url per recipient.
    """
    award_map: dict[str, dict] = {}
    award_order: list[str] = []
    for f in files:
        parts = split_name_parts(f["stem"])
        if parts and parts[0].isdigit():
            parts = parts[1:]  # strip leading order-number segment, e.g. "01"
        if len(parts) < 2:
            print(f"  WARNING: skipping award image '{f['name']}' - "
                  f"expected 'Award Title -- Name -- Designation'")
            continue
        title = parts[0]
        name  = parts[1] if len(parts) > 1 else ""
        desig = parts[2] if len(parts) > 2 else ""
        if title not in award_map:
            award_map[title] = {"title": title, "folder": "", "recipients": []}
            award_order.append(title)
        award_map[title]["recipients"].append({"name": name, "designation": desig, "image_url": f["url"]})
    return [award_map[t] for t in award_order]


def parse_new_joinees(files: list[dict]) -> list[dict]:
    """
    Parse filenames as 'Name--Designation.ext', returning a list of
    {name, designation, image_url}.
    """
    people = []
    for f in files:
        parts = split_name_parts(f["stem"])
        if len(parts) < 2:
            print(f"  WARNING: skipping new joinee image '{f['name']}' - "
                  f"expected 'Name -- Designation'")
            continue
        people.append({"name": parts[0], "designation": parts[1], "image_url": f["url"]})
    return people


def parse_named_images(files: list[dict]) -> list[dict]:
    """
    Customer logos, announcement banners and certificates all follow the same
    shape: any filename, the stem carried through as the name.
    Returns a list of {name, image_url}.
    """
    return [{"name": f["stem"], "image_url": f["url"]} for f in files]


# Words that appear in exported filenames without saying anything about which
# post the picture belongs to.
_IMAGE_NOISE = {
    "scaled", "copy", "final", "image", "img", "photo", "blog", "banner",
    "thumbnail", "thumb", "with", "and", "the", "for", "from", "a", "an",
    "in", "of", "to", "on", "at", "is", "it", "by", "www", "biztechcs", "com",
    "https", "http", "blogs",
}


def _words(text: str) -> set[str]:
    """Lowercase words worth comparing, from a URL slug or a filename."""
    return {
        w for w in re.split(r"[^a-z0-9]+", text.lower())
        if len(w) > 2 and w not in _IMAGE_NOISE and not w.isdigit()
    }


def match_blog_images(posts: list[dict], images: list[dict]) -> list[dict]:
    """
    Pair each blog post with the folder image whose filename best matches its
    URL slug. Returns the images reordered to line up with *posts*.

    Folder order is alphabetical, which has no reason to match the order the
    posts were pasted in: 'Go-Live-Checklist…' sorts after 'Data-Structure…'
    even though its post came first, so pairing by index silently put each
    title under the other post's picture — and the result looks completely
    normal, which is what makes it worth doing properly.

    Anything that can't be matched confidently keeps folder order, so this can
    only improve on the previous behaviour, never scramble a correct pairing.
    """
    if not posts or not images:
        return images

    remaining = list(images)
    matched: dict[int, dict] = {}

    # Best match first, so a strong pairing isn't stolen by a weaker one.
    scores = []
    for pi, post in enumerate(posts):
        target = _words(post.get("url", "") or post.get("title", ""))
        for image in images:
            overlap = len(target & _words(Path(image["name"]).stem))
            if overlap >= 2:
                scores.append((overlap, pi, image))
    scores.sort(key=lambda s: -s[0])

    for _overlap, pi, image in scores:
        if pi in matched or image not in remaining:
            continue
        matched[pi] = image
        remaining.remove(image)

    ordered = []
    for pi in range(len(posts)):
        if pi in matched:
            ordered.append(matched[pi])
        elif remaining:
            ordered.append(remaining.pop(0))
    ordered.extend(remaining)

    if matched:
        print(f"    blogs: matched {len(matched)} image(s) to posts by name")
    return ordered


def resolve_all_images(data: dict):
    """
    Walk all image URL fields in data and resolve folder/Drive/GitHub URLs
    to final direct image URLs. Modifies data in place.
    """
    img = data.setdefault("image_urls", {})

    print("  Resolving image URLs...")

    # Single-image sections — first file in the folder wins
    for key in ["ceo_photo", "campaign_banner", "hr_wellness_banner"]:
        raw = img.get(key, "")
        found = files_from_local_path(raw) or local_files(ROW_DATA_FOLDERS[key])
        if found:
            img[key] = found[0]["url"]
            print(f"    {key}: {found[0]['name']} (local)")
            continue
        img[key] = resolve_single_url(raw) if raw else ""
        print(f"    {key}: {img[key][:80] if img[key] else '(not found)'}")

    # Delivery Insights / Acknowledging Excellence — one logo per entry, taken
    # in filename order so the folder controls which entry gets which logo.
    for section, folder_key in (("delivery_insights", "delivery"),
                                ("acknowledging_excellence", "excellence")):
        entries = data.get(section, [])
        # A local folder path on any of the rows applies to the whole section.
        logos = next(
            (f for e in entries if (f := files_from_local_path(e.get("logo_url", "")))),
            [],
        ) or local_files(ROW_DATA_FOLDERS[folder_key])
        for i, entry in enumerate(entries):
            if i < len(logos):
                entry["logo_url"] = logos[i]["url"]
            elif entry.get("logo_url"):
                entry["logo_url"] = resolve_single_url(entry["logo_url"])
        source = f"{len(logos)} logo(s) from Row Data" if logos else "sheet links"
        print(f"    {section}: {len(entries)} entry(ies), {source}")

    # HR events — events/event1, events/event2, … matched to the sheet's event
    # rows by order.
    for n, event in enumerate(data.get("hr", {}).get("events", []), start=1):
        raws = event.pop("_photos_raw", [])
        found = next(
            (f for raw in raws if (f := files_from_local_path(raw))),
            [],
        ) or local_files(f"events/event{n}")
        if found:
            event["photos"] = [f["url"] for f in found]
            continue
        photos = []
        for raw in raws:
            photos.extend(resolve_folder_urls(raw))
        event["photos"] = photos
    n_events = len(data.get("hr", {}).get("events", []))
    n_event_photos = sum(len(e["photos"]) for e in data.get("hr", {}).get("events", []))
    print(f"    hr events: {n_events} event(s), {n_event_photos} photo(s) total")

    # Folder-driven sections. Each is filled from its Row Data folder; only if
    # that's empty does it fall back to a folder URL in the sheet, and failing
    # that to per-row image links.
    #
    # (data key, Row Data folder key, sheet-folder-URL key, parser, label)
    FOLDER_SECTIONS = [
        ("_awards",        "awards",         "_award_folder_urls",
         parse_awards,       "award(s)"),
        ("new_joinees",    "new_joinee",     "_new_joinee_folder_urls",
         parse_new_joinees,  "new joinee(s)"),
        ("new_customers",  "new_customer",   "_new_customer_folder_urls",
         parse_named_images, "customer logo(s)"),
        ("announcements",  "announcement",   "_announcement_folder_urls",
         parse_named_images, "announcement(s)"),
        ("certifications", "certification",  "_certification_folder_urls",
         parse_named_images, "certificate(s)"),
    ]

    for data_key, folder_key, sheet_key, parser, label in FOLDER_SECTIONS:
        # Awards live one level down, under rewards.
        container = data["rewards"] if data_key == "_awards" else data
        key = "awards" if data_key == "_awards" else data_key

        sheet_folders = data.pop(sheet_key, [])

        # A local folder path in the sheet wins; otherwise the conventional one.
        found = next(
            (f for raw in sheet_folders if (f := files_from_local_path(raw))),
            [],
        ) or local_files(ROW_DATA_FOLDERS[folder_key])
        if found:
            parsed = parser(found)
            if parsed:
                container[key] = parsed
                print(f"    {key}: {len(parsed)} {label} from local folder")
                continue

        matched = False
        for folder_url in sheet_folders:
            parsed = parser(github_folder_files(folder_url))
            if parsed:
                container[key] = parsed
                print(f"    {key}: {len(parsed)} {label} from GitHub")
                matched = True
                break
        if matched:
            continue

        # No folder either side — resolve whatever per-row links exist.
        if key == "awards":
            rows = container[key]
            for award in rows:
                for r in award["recipients"]:
                    if r.get("image_url"):
                        r["image_url"] = resolve_single_url(r["image_url"])
        else:
            rows = container.get(key, [])
            for row in rows:
                if row.get("image_url"):
                    row["image_url"] = resolve_single_url(row["image_url"])
        if rows:
            print(f"    {key}: {len(rows)} {label} from sheet rows")
        elif ROW_DATA_DIR is not None:
            print(f"    {key}: none - add images to Row Data/{ROW_DATA_FOLDERS[folder_key]}")

    # Photo galleries — Expo / Training / Workshop, each just a folder.
    for key, _field, folder, label in GALLERY_SECTIONS:
        gallery = data.setdefault(key, {"title": "", "photos": []})
        raw = gallery.pop("_photos_raw", "")
        found = files_from_local_path(raw) or local_files(folder)
        if found:
            gallery["photos"] = [f["url"] for f in found]
        elif raw:
            gallery["photos"] = resolve_folder_urls(raw)
        print(f"    {key}: {len(gallery['photos'])} photo(s)"
              f"{' - ' + gallery['title'] if gallery['title'] else ''}")

    # Blog images — matched to posts by name, not folder order.
    posts = data["marketing"]["blog_posts"]
    blog_images = next(
        (f for p in posts if (f := files_from_local_path(p.get("image_url", "")))),
        [],
    ) or local_files(ROW_DATA_FOLDERS["blogs"])
    blog_images = match_blog_images(posts, blog_images)

    # Otherwise: if multiple blogs share a folder URL, assign images by index
    folder_cache: dict[str, list[str]] = {}
    folder_counter: dict[str, int] = {}
    for i, post in enumerate(data["marketing"]["blog_posts"], 1):
        if i <= len(blog_images):
            post["image_url"] = blog_images[i - 1]["url"]
            print(f"    blog{i}: Row Data/{ROW_DATA_FOLDERS['blogs']}/{blog_images[i - 1]['name']}")
            continue
        raw = post.get("image_url", "")
        if not raw:
            print(f"    blog{i}: (no image)")
            continue
        info = parse_github_url(raw)
        if info and info["is_folder"]:
            if raw not in folder_cache:
                files = list_github_folder(info["owner"], info["repo"], info["branch"], info["path"])
                folder_cache[raw] = [f["url"] for f in files]
                folder_counter[raw] = 0
            idx = folder_counter[raw]
            urls = folder_cache[raw]
            post["image_url"] = urls[idx] if idx < len(urls) else ""
            folder_counter[raw] = idx + 1
        else:
            post["image_url"] = resolve_single_url(raw)
        print(f"    blog{i}: {post['image_url'][:80] if post['image_url'] else '(not found)'}")


def download_drive_xlsx(drive_url: str) -> openpyxl.Workbook:
    """Download an Excel file from Google Drive (must be shared publicly)."""
    fid = drive_file_id(drive_url)
    if not fid:
        raise ValueError(f"Could not extract file ID from: {drive_url}")
    url = f"https://drive.google.com/uc?export=download&id={fid}"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req) as resp:
        return openpyxl.load_workbook(io.BytesIO(resp.read()))


# ── Anniversary parser ────────────────────────────────────────────────────

def parse_anniversaries(wb: openpyxl.Workbook, sheet_name: str) -> list:
    """
    Find the sheet named sheet_name (e.g. 'April-2026') in wb.
    Each sheet has two tables side by side:
      - Birthday table   (columns A–G)
      - Work Anniversary table (columns I onwards)
    Row 1: table title headers ('Birthday' | 'Work Anniversary')
    Row 2: column headers (E.Id | Employee Name | ... | Tenure | ...)
    Row 3+: data rows

    Extracts Employee Name + Tenure from the Work Anniversary table only,
    groups names by years, returns list sorted by years descending.
    """
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Using sheet: '{sheet_name}'")
    else:
        # Partial match fallback on the month part
        month_part = sheet_name.split("-")[0].lower()
        ws = None
        for name in wb.sheetnames:
            if month_part in name.lower():
                ws = wb[name]
                print(f"  Sheet '{sheet_name}' not found — using '{name}'")
                break
        if not ws:
            print(f"  WARNING: No sheet found for '{sheet_name}'. Available: {wb.sheetnames}")
            return []

    # ── Find where 'Work Anniversary' header is (row 1) ──────────────────
    # The sheet layout: Birthday header in col A, Work Anniversary header in col I
    wa_col_start = None
    title_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for col_idx, cell in enumerate(title_row, start=1):
        if cell and "work anniversary" in str(cell).lower():
            wa_col_start = col_idx
            break

    if not wa_col_start:
        print("  WARNING: Could not find 'Work Anniversary' header in row 1.")
        return []

    # ── Find Employee Name & Tenure columns in row 2, right of wa_col_start ─
    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    name_col = tenure_col = None

    for col_idx, cell in enumerate(header_row, start=1):
        if col_idx < wa_col_start:
            continue          # skip Birthday table columns
        text = str(cell).lower().strip() if cell else ""
        if "employee name" in text and name_col is None:
            name_col = col_idx
        if text == "tenure" and tenure_col is None:
            tenure_col = col_idx

    if not name_col or not tenure_col:
        print(f"  WARNING: Could not locate Employee Name / Tenure in Work Anniversary table.")
        return []

    print(f"  Work Anniversary starts at col {wa_col_start} | name_col={name_col} | tenure_col={tenure_col}")

    # ── Read data rows (row 3 onwards) ────────────────────────────────────
    year_map: dict[int, list[str]] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        name   = cell_str(row[name_col - 1])   if name_col   <= len(row) else ""
        tenure = cell_str(row[tenure_col - 1]) if tenure_col <= len(row) else ""
        if not name or not tenure:
            continue
        years = parse_years(tenure)
        if years == 0:
            continue
        year_map.setdefault(years, []).append(name)

    return [{"years": y, "names": names} for y, names in sorted(year_map.items(), reverse=True)]


def parse_anniversaries_from_cell(raw: str) -> list:
    """
    Parse anniversary data pasted straight into the sheet's Extra column,
    one person per line as 'Name<TAB>N years':

        Viral savaj      13 years
        Dhaval Panara    10 years

    Pasting a block from another spreadsheet gives exactly this, so it saves
    maintaining a separate linked workbook. Comma and semicolon also work as
    the separator. Returns the same {years, names} grouping as the workbook
    parser, so the rest of the pipeline can't tell the difference.
    """
    year_map: dict[int, list[str]] = {}
    for line in str(raw or "").splitlines():
        line = line.strip()
        if not line:
            continue
        for sep in ("\t", ";", ","):
            if sep in line:
                name, _, tenure = line.partition(sep)
                break
        else:
            # No separator — fall back to splitting off a trailing "12 years".
            m = re.match(r"^(.*?)\s+(\d+\s*years?)$", line, re.IGNORECASE)
            if not m:
                print(f"  WARNING: skipping anniversary line (no name/tenure separator): {line!r}")
                continue
            name, tenure = m.group(1), m.group(2)
        name, years = name.strip(), parse_years(tenure)
        if not name or years == 0:
            print(f"  WARNING: skipping anniversary line (no name or year count): {line!r}")
            continue
        year_map.setdefault(years, []).append(name)
    return [{"years": y, "names": names} for y, names in sorted(year_map.items(), reverse=True)]


def parse_blogs_from_cell(raw: str) -> list[dict]:
    """
    Parse blog posts pasted into a single cell, one per line as
    'Title<TAB>https://…'. The master sheet has one Marketing Highlights row
    for all of them, rather than a numbered row per post.

    A bare URL on its own line is accepted too; the title then falls back to
    the URL's last path segment so the card is never left blank.
    """
    posts = []
    for line in str(raw or "").splitlines():
        line = line.strip()
        if not line:
            continue
        title, url = line, ""
        for sep in ("\t", "|", ";"):
            if sep in line:
                title, _, url = line.partition(sep)
                break
        else:
            # No separator: either a bare URL, or "Title http://…".
            m = re.match(r"^(.*?)\s*(https?://\S+)$", line)
            if m:
                title, url = m.group(1), m.group(2)
            elif line.startswith("http"):
                title, url = "", line
        title, url = title.strip(), url.strip()
        if not url and title.startswith("http"):
            title, url = "", title
        if not url:
            print(f"  WARNING: skipping blog line (no link): {line!r}")
            continue
        if not title:
            title = url.rstrip("/").rsplit("/", 1)[-1].replace("-", " ").title()
        posts.append({"title": title, "url": url, "image_url": ""})
    return posts


def parse_openings_from_cell(raw: str) -> list[dict]:
    """
    Parse open positions pasted into the sheet's Extra column, one per line:

        Position                        Experience   Opening
        SEO Executive                   Fresher      1
        Business Development Executive  2+ years     2

    A header line is detected and used to map the columns, so Position /
    Experience / Opening can be in any order; without one, that order is
    assumed. Tab, semicolon and comma all work as the separator.
    """
    rows = []
    order = ["title", "experience", "openings"]
    for line in str(raw or "").splitlines():
        line = line.strip()
        if not line:
            continue
        for sep in ("\t", ";", ","):
            if sep in line:
                parts = [p.strip() for p in line.split(sep)]
                break
        else:
            parts = [line]

        lowered = [p.lower() for p in parts]
        if any(h in lowered for h in ("position", "designation", "role")):
            # Header row — remember the column order, don't emit a position.
            mapping = []
            for p in lowered:
                if p in ("position", "designation", "role"):   mapping.append("title")
                elif p.startswith("exp"):                      mapping.append("experience")
                elif p.startswith("open") or "count" in p or "vacan" in p:
                    mapping.append("openings")
                else:                                          mapping.append("")
            if "title" in mapping:
                order = mapping
            continue

        row = {"title": "", "experience": "", "openings": 1}
        for i, part in enumerate(parts):
            field = order[i] if i < len(order) else ""
            if field == "openings":
                try:
                    row["openings"] = int(float(part)) if part else 1
                except ValueError:
                    row["openings"] = 1
            elif field:
                row[field] = part
        if not row["title"]:
            print(f"  WARNING: skipping openings line (no position): {line!r}")
            continue
        rows.append(row)
    return rows


# ── New Openings (external sheet) parser ────────────────────────────────────

def parse_openings_sheet(wb: openpyxl.Workbook, month: str) -> list[dict]:
    """
    Parse an external 'New Openings' sheet, columns: Month | Position | Experience | Opening.
    Row 1: title, Row 2: headers, Row 3+: data — Month is only filled on the
    first row of each month's block, so it's forward-filled down.
    Returns positions for the given month only.
    """
    ws = wb[wb.sheetnames[0]]
    positions = []
    current_month = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        month_cell = cell_str(row[0]) if len(row) > 0 else ""
        if month_cell:
            current_month = month_cell
        if not current_month or current_month.strip().lower() != (month or "").strip().lower():
            continue
        title = cell_str(row[1]) if len(row) > 1 else ""
        if not title:
            continue
        experience = cell_str(row[2]) if len(row) > 2 else ""
        opening_raw = row[3] if len(row) > 3 else None
        try:
            count = int(float(opening_raw)) if opening_raw not in (None, "") else 1
        except (ValueError, TypeError):
            count = 1
        positions.append({"title": title, "openings": count, "experience": experience})
    return positions


# ── Main content sheet parser ─────────────────────────────────────────────

def parse_content_sheet(wb: openpyxl.Workbook) -> dict:
    """Parse the Newsletter Content sheet into the content.json structure."""

    sheet_name = "Newsletter Content" if "Newsletter Content" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    data = {
        "month": None,
        "year":  None,
        "ceo_desk": {},
        "delivery_insights": [],
        "acknowledging_excellence": [],
        "hr": {"events": []},
        "rewards": {"awards": []},
        "anniversaries": [],
        "new_joinees": [],
        "new_customers": [],
        "announcements": [],
        "certifications": [],
        "new_openings": {
            "positions": [],
            "view_all_url": "https://www.biztechcs.com/career/",
        },
        "marketing": {
            "blog_posts": [],
            "view_all_url": "https://www.biztechcs.com/blog/",
        },
        "image_urls": {
            "ceo_photo": "",
            "campaign_banner": "",
            "hr_wellness_banner": "",
        },
        "footer": FOOTER,
        # Title + photos each; see GALLERY_SECTIONS.
        "expo":     {"title": "", "photos": []},
        "training": {"title": "", "photos": []},
        "workshop": {"title": "", "photos": []},
        # Sections explicitly marked NA in the sheet; consumed by generate.py.
        "_na_sections": set(),
    }

    award_map: dict = {}   # award title → award dict
    _legacy_delivery: dict = {}    # merges old single-row delivery_description/delivery_logo fields
    _legacy_excellence: dict = {}  # merges old single-row excellence_client/_testimonial/_logo fields
    _legacy_hr_event_title = ""    # merges old single-event hr_event_title field
    _legacy_event_photos_raw: list = []  # merges old single-event event_photo rows

    # The master sheet everyone fills in reads
    #   Sr No. | Newsletter Sections | POC | Content | Photos Folder | Field
    # so the machine-readable bits sit to the right of the human-readable ones.
    # Detected by its header, and remapped onto the same (field, value, folder)
    # the original layout uses, so one parser serves both.
    master = None
    # Scanned generously: the master sheet carries a block of instructions
    # above its header, and people add to it.
    for probe in ws.iter_rows(min_row=1, max_row=25):
        headers = {cell_str(c.value).lower(): c.column - 1 for c in probe if c.value}
        if "newsletter sections" in headers and "field" in headers:
            master = {
                "field":   headers["field"],
                "content": headers.get("content", 3),
                "folder":  headers.get("photos folder", 4),
                "header_row": probe[0].row,
            }
            print(f"  Master-format sheet detected (header row {master['header_row']}).")
            break

    for row in ws.iter_rows():  # Cell objects, not values_only — the Drive Link
                                 # column needs .hyperlink (see below), since a
                                 # cell inserted as "Insert link" with custom
                                 # display text stores the URL only there, not
                                 # in .value.
        def at(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        if master:
            if row[0].row <= master["header_row"]:
                continue
            field_cell   = at(master["field"])
            content_cell = at(master["content"])
            folder_cell  = at(master["folder"])
            field = cell_str(field_cell.value) if field_cell else ""
            # Content carries both the plain value and any pasted block, so
            # multi-line entries (anniversaries, openings, blogs) work here too.
            row = [field_cell, content_cell, content_cell, None, folder_cell]
        else:
            field = cell_str(row[0].value) if row else ""

        # Skip blank rows, section dividers, legend row
        if not field or field.startswith("▸") or field.lower() == "field":
            continue

        # blank_if_na: a cell reading "NA" (or "-", "none", …) means nothing
        # this month, so it is treated exactly as if it were empty and the
        # section drops out rather than rendering the word as content.
        # An explicit NA anywhere on the row removes the section outright.
        if FIELD_SECTION.get(field) and any(
            is_na(c.value) for c in row[1:5] if c is not None
        ):
            data["_na_sections"].add(FIELD_SECTION[field])

        def cell_at(idx):
            cell = row[idx] if idx < len(row) else None
            return blank_if_na(cell.value) if cell is not None else None

        val    = cell_at(1)
        extra  = cell_at(2)
        extra2 = cell_at(3)
        drive_cell = row[4] if len(row) > 4 else None
        drive = blank_if_na(drive_cell.value) if drive_cell is not None else None
        if (
            not (isinstance(drive, str) and drive.strip().lower().startswith("http"))
            and drive_cell is not None
            and drive_cell.hyperlink
            and drive_cell.hyperlink.target
        ):
            # The cell's own text isn't a usable URL — fall back to the
            # hyperlink target, for a cell inserted as a rich link with a
            # custom display label (e.g. Google Sheets "Insert link").
            # Only used as a fallback: when the visible text IS already a
            # URL, trust it over the hyperlink — copy-pasting a row and
            # editing its visible URL text can leave the underlying
            # hyperlink target stale (still pointing at the original row's
            # link), which would otherwise silently override correct data.
            drive = drive_cell.hyperlink.target

        # ── General ──────────────────────────────────────────────────────
        if   field == "month":        data["month"] = cell_str(val)
        elif field == "year":         data["year"]  = str(int(float(str(val)))) if val else None
        elif field == "ceo_name":     data["ceo_desk"]["ceo_name"]  = cell_str(val)
        elif field == "ceo_title":    data["ceo_desk"]["ceo_title"] = cell_str(val)
        elif field == "ceo_message":  data["ceo_desk"]["message"]   = cell_str(val)

        # ── Image rows — store raw URL, resolved later ───────────────────
        elif field == "ceo_photo":
            data["image_urls"]["ceo_photo"] = cell_str(drive)
        elif field == "campaign_banner":
            data["image_urls"]["campaign_banner"] = cell_str(drive)
        elif field == "hr_wellness_banner":
            data["image_urls"]["hr_wellness_banner"] = cell_str(drive)
            if cell_str(val):
                data["hr"]["wellness_title"] = cell_str(val)
        elif field == "event_photo":   # legacy single-event rows
            raw = cell_str(drive)
            if raw:
                _legacy_event_photos_raw.append(raw)

        # ── Delivery Insights (repeatable: one row per entry) ─────────────
        elif field == "delivery":
            data["delivery_insights"].append({
                "description": cell_str(val),
                "logo_url":    cell_str(drive),   # resolved later
            })
        elif field == "delivery_description":   # legacy single-entry fields
            _legacy_delivery["description"] = cell_str(val)
        elif field == "delivery_logo":
            _legacy_delivery["logo_url"] = cell_str(drive)

        # ── Acknowledging Excellence (repeatable: one row per entry) ──────
        elif field == "excellence":
            data["acknowledging_excellence"].append({
                "client_label": cell_str(extra),
                "testimonial":  cell_str(val),
                "logo_url":     cell_str(drive),   # resolved later
            })
        elif field == "excellence_client":       # legacy single-entry fields
            _legacy_excellence["client_label"] = cell_str(val)
        elif field == "excellence_testimonial":
            _legacy_excellence["testimonial"] = cell_str(val)
        elif field == "excellence_logo":
            _legacy_excellence["logo_url"] = cell_str(drive)

        # ── HR Insider ────────────────────────────────────────────────────
        elif field == "hr_wellness_title":
            data["hr"]["wellness_title"] = cell_str(val)
        elif re.match(r"^hr_event\d*$", field):   # repeatable: 1 row per event (Value=title, Drive Link=photo folder); accepts hr_event, hr_event1, hr_event2, ...
            title = cell_str(val)
            raw   = cell_str(drive)
            if title or raw:
                data["hr"]["events"].append({"title": title, "_photos_raw": [raw] if raw else []})
        elif field == "hr_event_title":   # legacy single-event field
            _legacy_hr_event_title = cell_str(val)

        # ── Awards ────────────────────────────────────────────────────────
        elif field == "award":
            title = cell_str(val)
            raw   = cell_str(drive)
            if not title:
                # Folder-based: images named "Title--Name--Designation.jpg"
                if raw and raw not in data.get("_award_folder_urls", []):
                    data.setdefault("_award_folder_urls", []).append(raw)
            else:
                # Explicit: title/name/designation in columns
                name  = cell_str(extra) or None
                desig = cell_str(extra2)
                if title not in award_map:
                    award_map[title] = {"title": title, "folder": "", "recipients": []}
                if name:
                    award_map[title]["recipients"].append({
                        "name":        name,
                        "designation": desig,
                        "image_url":   raw,   # resolved later
                    })

        # ── Employee Certifications ────────────────────────────────────────
        elif field == "employee_certification":
            title = cell_str(val)
            raw   = cell_str(drive)
            if not title:
                # Folder-based: any certificate image files, no naming convention required
                if raw and raw not in data.get("_certification_folder_urls", []):
                    data.setdefault("_certification_folder_urls", []).append(raw)
            else:
                # Explicit row
                data["certifications"].append({
                    "name":      title,
                    "image_url": raw,   # resolved later
                })

        # ── New Additions (new joinees) ──────────────────────────────────
        elif field == "new_joinee":
            name = cell_str(val)
            raw  = cell_str(drive)
            if not name:
                # Folder-based: images named "Name--Designation.jpg"
                if raw and raw not in data.get("_new_joinee_folder_urls", []):
                    data.setdefault("_new_joinee_folder_urls", []).append(raw)
            else:
                # Explicit row
                data["new_joinees"].append({
                    "name":        name,
                    "designation": cell_str(extra),
                    "image_url":   raw,   # resolved later
                })

        # ── Newly Added Customers ─────────────────────────────────────────
        elif field == "new_customer":
            name = cell_str(val)
            raw  = cell_str(drive)
            if not name:
                # Folder-based: any logo image files, no naming convention required
                if raw and raw not in data.get("_new_customer_folder_urls", []):
                    data.setdefault("_new_customer_folder_urls", []).append(raw)
            else:
                # Explicit row
                data["new_customers"].append({
                    "name":      name,
                    "image_url": raw,   # resolved later
                })

        # ── Upcoming Event Announcements ──────────────────────────────────
        elif field == "new_announcement":
            title = cell_str(val)
            raw   = cell_str(drive)
            if not title:
                # Folder-based: any banner image files, no naming convention required
                if raw and raw not in data.get("_announcement_folder_urls", []):
                    data.setdefault("_announcement_folder_urls", []).append(raw)
            else:
                # Explicit row
                data["announcements"].append({
                    "name":      title,
                    "image_url": raw,   # resolved later
                })

        # ── New Openings ──────────────────────────────────────────────────
        elif field == "opening":
            if val:
                try:
                    count = int(float(str(extra))) if extra else 1
                except ValueError:
                    count = 1
                data["new_openings"]["positions"].append({
                    "title":      cell_str(val),
                    "openings":   count,
                    "experience": cell_str(extra2),
                })
        elif field == "new_openings_sheet":
            # External sheet link — Drive Link column = Google Sheets URL.
            # When present, its data replaces any "opening" rows above.
            url = cell_str(drive)
            if url and url.startswith("http"):
                data["_openings_sheet_url"] = url
            # Positions pasted straight into the Extra column, one per line.
            if cell_str(extra):
                data["_openings_inline"] = cell_str(extra)

        # ── Marketing Blogs (blog1, blog2, …) ────────────────────────────
        elif re.match(r"^blog\d*$", field) or field == "marketing_highlights":
            text = cell_str(val)
            if "\n" in text or field == "marketing_highlights":
                # One cell holding every post, "Title<TAB>URL" per line — the
                # shape the master sheet uses.
                for post in parse_blogs_from_cell(text):
                    post["image_url"] = cell_str(drive)   # resolved later
                    data["marketing"]["blog_posts"].append(post)
            elif text:
                data["marketing"]["blog_posts"].append({
                    "title":     text,
                    "url":       cell_str(extra),
                    "image_url": cell_str(drive),   # resolved later
                })

        # ── Photo galleries (Expo / Training / Workshop) ──────────────────
        elif field in {f for _k, f, _d, _l in GALLERY_SECTIONS}:
            key = next(k for k, f, _d, _l in GALLERY_SECTIONS if f == field)
            data[key]["title"] = cell_str(val)
            if cell_str(drive):
                data[key]["_photos_raw"] = cell_str(drive)

        # ── Anniversary — tab name in Value, Drive file in Drive Link ───────
        elif field == "anniversary":
            # Value may be a date object (Excel stores "March-2026" as a date)
            if hasattr(val, "strftime"):
                tab = val.strftime("%B-%Y")   # datetime → "March-2026"
            else:
                tab = cell_str(val)
            url = cell_str(drive)             # Drive link to the xlsx file
            if tab:
                data["_anniv_tab"] = tab
            if url and url.startswith("http"):
                data["_anniv_drive_url"] = url
            # Names pasted straight into the Extra column, one per line as
            # "Name<TAB>N years" — used in preference to the linked workbook.
            if cell_str(extra):
                data["_anniv_inline"] = cell_str(extra)

        # Unknown fields: silently skip

    data["rewards"]["awards"] = list(award_map.values())

    if not data["delivery_insights"] and _legacy_delivery:
        data["delivery_insights"].append(_legacy_delivery)
    if not data["acknowledging_excellence"] and _legacy_excellence:
        data["acknowledging_excellence"].append(_legacy_excellence)
    if not data["hr"]["events"] and (_legacy_hr_event_title or _legacy_event_photos_raw):
        data["hr"]["events"].append({"title": _legacy_hr_event_title, "_photos_raw": _legacy_event_photos_raw})

    # A set can't be written to JSON.
    data["_na_sections"] = sorted(data["_na_sections"])

    return data


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    # Keep this script's own prints interleaved correctly with the generate.py /
    # screenshot_sections.py subprocess output below (stdout is block-buffered,
    # not line-buffered, when piped instead of attached to a real terminal).
    sys.stdout.reconfigure(line_buffering=True)

    # ── Step 1: Newsletter content ────────────────────────────────────────
    # --sheet=<file.xlsx> drives the run from a local Excel file instead of the
    # shared Google Sheet — useful when you only have view access to the sheet,
    # or want to prepare a month without touching the live copy.
    local_sheet = next(
        (a.split("=", 1)[1] for a in sys.argv if a.startswith("--sheet=")), None
    )
    if local_sheet:
        sheet_path = Path(local_sheet)
        if not sheet_path.is_file():
            raise SystemExit(f"--sheet file not found: {sheet_path}")
        print(f"Reading Newsletter Content from local file: {sheet_path}")
        content_wb = openpyxl.load_workbook(sheet_path)
    else:
        print("Downloading Newsletter Content sheet...")
        content_wb = download_xlsx(CONTENT_SHEET_ID)
    data = parse_content_sheet(content_wb)
    print(f"  month={data['month']}, year={data['year']}")
    print(f"  Awards : {[a['title'] for a in data['rewards']['awards']]}")
    print(f"  New Joinees: {len(data['new_joinees'])}")
    print(f"  New Customers: {len(data['new_customers'])}")
    print(f"  Announcements: {len(data['announcements'])}")
    print(f"  Certifications: {len(data['certifications'])}")
    print(f"  Blogs   : {len(data['marketing']['blog_posts'])}")

    # ── Step 1b: New Openings — external sheet link overrides sheet rows ───
    openings_sheet_url = data.pop("_openings_sheet_url", None)
    openings_inline = data.pop("_openings_inline", None)
    if openings_inline:
        # Positions typed into the sheet win — no external workbook needed.
        print("Reading New Openings from the sheet...")
        positions = parse_openings_from_cell(openings_inline)
        if positions:
            data["new_openings"]["positions"] = positions
    elif openings_sheet_url:
        print(f"Downloading New Openings sheet...")
        try:
            openings_sheet_id = drive_file_id(openings_sheet_url)
            openings_wb = download_xlsx(openings_sheet_id)
            positions = parse_openings_sheet(openings_wb, data["month"])
            if positions:
                data["new_openings"]["positions"] = positions
            else:
                print(f"  WARNING: No openings found for month '{data['month']}' in external sheet.")
        except Exception as e:
            print(f"  WARNING: Could not load New Openings sheet ({e}); using sheet-row data instead.")
    print(f"  Openings: {len(data['new_openings']['positions'])}")

    # ── Step 2: Work Anniversary data ────────────────────────────────────
    anniv_tab       = data.pop("_anniv_tab",       None)
    anniv_drive_url = data.pop("_anniv_drive_url", None)
    anniv_inline    = data.pop("_anniv_inline",    None)

    if anniv_inline:
        # Names typed straight into the sheet win — no download, no linked
        # workbook, nothing else to keep in sync.
        print("\nReading Work Anniversary data from the sheet...")
        anniversaries = parse_anniversaries_from_cell(anniv_inline)
    else:
        # Sheet name to look up — from content sheet row, else from month/year
        sheet_name = anniv_tab or f"{data['month']}-{data['year']}"

        print(f"\nDownloading Work Anniversary file (tab: '{sheet_name}')...")
        if anniv_drive_url:
            try:
                anniv_wb = download_drive_xlsx(anniv_drive_url)
                print("  Downloaded from Drive link in sheet.")
            except Exception as e:
                print(f"  Drive download failed ({e}), falling back to Sheet ID...")
                anniv_wb = load_xlsx(ANNIV_SHEET_ID, "anniversaries.xlsx")
        else:
            anniv_wb = load_xlsx(ANNIV_SHEET_ID, "anniversaries.xlsx")

        anniversaries = parse_anniversaries(anniv_wb, sheet_name)

    data["anniversaries"] = anniversaries

    if anniversaries:
        print(f"  {len(anniversaries)} anniversary group(s):")
        for a in anniversaries:
            label = "year" if a["years"] == 1 else "years"
            print(f"    {a['years']} {label}: {', '.join(a['names'])}")
    else:
        print("  No anniversary data found.")

    # ── Step 3: Resolve all image URLs (local Row Data → GitHub folders) ───
    global ROW_DATA_DIR
    ROW_DATA_DIR = BASE_DIR / f"{data['month']}-{data['year']}" / "Row Data"
    print("\nResolving images...")
    if ROW_DATA_DIR.is_dir():
        have = sorted(p.name for p in ROW_DATA_DIR.iterdir() if p.is_dir())
        print(f"  Local source folder: {ROW_DATA_DIR}")
        print(f"    {len(have)} subfolder(s): {', '.join(have) if have else '(empty - falling back to GitHub)'}")
    else:
        print(f"  No local folder at {ROW_DATA_DIR} - using the GitHub folders from the sheet.")
    resolve_all_images(data)

    # ── Step 4: Write content.json ────────────────────────────────────────
    OUTPUT.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\nDone -> {OUTPUT}")

    # ── Step 5: Render the browser-preview HTML for review ────────────────
    # This stops at the preview on purpose: the section PNGs are what actually
    # get mailed, so they're only worth regenerating once the content in the
    # preview has been approved. Pass --all to run the whole chain in one go.
    print("\nRendering updated newsletter for review...")
    subprocess.run([sys.executable, str(BASE_DIR / "generate.py")], check=True)

    run_all = "--all" in sys.argv
    if run_all:
        subprocess.run([sys.executable, str(BASE_DIR / "screenshot_sections.py")], check=True)
        subprocess.run([sys.executable, str(BASE_DIR / "generate_simple_email.py")], check=True)
    else:
        print("\n" + "-" * 70)
        print("REVIEW the preview above, then run:")
        print(f"  1. python screenshot_sections.py     -> PNGs in {data['month']}-{data['year']}/Section wise images")
        print("  2. upload those PNGs                 -> see UPLOAD_ME.txt in that folder")
        print("  3. python generate_simple_email.py   -> final Gmail/Outlook-safe HTML")
        print("\n(or 'python import_content.py --all' to skip the review gate)")
        print("-" * 70)


if __name__ == "__main__":
    main()
