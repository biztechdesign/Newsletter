"""
migrate_to_master.py
====================
Rewrites an old-format content sheet into the shared master format, keeping
whatever content is already in it.

The old sheet is a flat Field/Value/Extra/Extra2/Drive Link list with section
divider rows and a legend. The master sheet is the one contributors actually
see: Sr No. / Newsletter Sections / POC / Content, plus the two wiring columns.
This carries the data across so a month already in progress doesn't have to be
retyped.

Usage:
    python migrate_to_master.py July-2026/content.xlsx

Writes the converted file in place and keeps the original as *.old.xlsx.
"""

import shutil
import sys
from pathlib import Path

import openpyxl

from create_master_template import build, ensure_folders

BASE_DIR = Path(__file__).parent

# Old field -> master field, where the name changed.
RENAMED = {
    "delivery_description": "delivery",
    "hr_event": "hr_event1",
    "blog1": "marketing_highlights",
    "blog2": "marketing_highlights",
}

# Old rows that carry no content of their own — a folder link, or a value the
# master sheet derives from somewhere else.
DROPPED = {"ceo_photo", "delivery_logo", "excellence_logo", "event_photo"}


def read_master(ws) -> tuple[dict, dict]:
    """
    ({field: [content, ...]}, {field: note}) from a sheet already in master
    format. Lets this be re-run without wiping what is already filled in —
    the alternative is a silent, total data loss on an accidental second run.
    """
    header_row = None
    cols = {}
    for probe in ws.iter_rows(min_row=1, max_row=25):
        headers = {str(c.value).strip().lower(): c.column for c in probe if c.value}
        if "newsletter sections" in headers and "field" in headers:
            header_row, cols = probe[0].row, headers
            break
    if header_row is None:
        return {}, {}

    content_col = cols.get("content", 4)
    field_col = cols["field"]
    note_col = cols.get("note")
    found: dict[str, list[str]] = {}
    notes: dict[str, str] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        field = str(ws.cell(row=r, column=field_col).value or "").strip()
        if not field:
            continue
        value = ws.cell(row=r, column=content_col).value
        text = str(value).strip() if value is not None else ""
        if text:
            found.setdefault(field, []).append(text)
        if note_col:
            note = ws.cell(row=r, column=note_col).value
            if note is not None and str(note).strip():
                notes[field] = str(note).strip()
    return found, notes


def read_old(path: Path) -> dict:
    """{field: [content, ...]} from an old-format sheet, in row order."""
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    found: dict[str, list[str]] = {}
    for row in ws.iter_rows():
        field = str(row[0].value).strip() if row[0].value else ""
        if not field or field.startswith("▸") or field.lower() == "field":
            continue
        if field in DROPPED:
            continue
        field = RENAMED.get(field, field)

        # Value first, then Extra — pasted blocks (anniversaries, openings)
        # live in Extra, everything else in Value.
        parts = []
        for cell in row[1:4]:
            text = str(cell.value).strip() if cell is not None and cell.value is not None else ""
            if text and not text.lower().startswith("row data"):
                parts.append(text)
        if field == "anniversary":
            # The old sheet put the tab-selector date in Value and the pasted
            # names in Extra. The master sheet takes the month from its own
            # month/year rows, so keep only the names.
            parts = [p for p in parts if "\n" in p or "\t" in p]
        if field == "marketing_highlights":
            # blog1/blog2 were Title in Value and URL in Extra; the master
            # sheet wants one "Title<TAB>URL" line per post.
            if len(parts) >= 2:
                parts = ["\t".join(parts[:2])]
            elif not parts:
                continue
        content = parts[0] if len(parts) == 1 else "\n".join(parts)
        if not content:
            continue
        found.setdefault(field, []).append(content)
    return found


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python migrate_to_master.py <sheet.xlsx>")
    target = Path(sys.argv[1])
    if not target.is_absolute():
        target = BASE_DIR / target
    if not target.is_file():
        raise SystemExit(f"Not found: {target}")

    wb_in = openpyxl.load_workbook(target)
    already, notes = read_master(wb_in[wb_in.sheetnames[0]])
    if already or notes:
        print("  Sheet is already in master format - refreshing it and keeping "
              "the Content and Note already filled in.")
        old = already
    else:
        old = read_old(target)
        notes = {}

    backup = target.with_suffix(".old.xlsx")
    shutil.copy(target, backup)

    build(target)   # lay down a fresh master sheet
    wb = openpyxl.load_workbook(target)
    ws = wb[wb.sheetnames[0]]

    header_row = next(
        r for r in range(1, ws.max_row + 1)
        if str(ws.cell(row=r, column=1).value or "").strip() == "Sr No."
    )

    # Fields appearing on more than one master row (excellence, hr_event…) are
    # filled in order, so a second testimonial lands on the second row.
    used: dict[str, int] = {}
    carried, missing = [], []
    for r in range(header_row + 1, ws.max_row + 1):
        field = str(ws.cell(row=r, column=6).value or "").strip()
        if not field or field not in old:
            continue
        values = old[field]
        i = used.get(field, 0)
        if i >= len(values):
            continue
        ws.cell(row=r, column=4).value = values[i]
        used[field] = i + 1
        carried.append(f"{field}: {values[i].splitlines()[0][:46]}")

    # Notes are per-field, so they survive a refresh alongside the content.
    note_col = len(ws[header_row])
    for r in range(header_row + 1, ws.max_row + 1):
        field = str(ws.cell(row=r, column=6).value or "").strip()
        if field in notes:
            ws.cell(row=r, column=7).value = notes[field]

    for field, values in old.items():
        left = len(values) - used.get(field, 0)
        if left > 0:
            missing.append(f"{field} x{left}")

    wb.save(target)
    made = ensure_folders(target)

    print(f"Migrated: {target}")
    print(f"  Backup:  {backup.name}")
    if made:
        print(f"  Created {len(made)} photo folder(s): {', '.join(made)}")
    print(f"\n  Carried across ({len(carried)}):")
    for line in carried:
        print(f"    {line}")
    if missing:
        print(f"\n  NOT carried — no matching row in the master sheet:")
        for line in missing:
            print(f"    {line}")


if __name__ == "__main__":
    main()
