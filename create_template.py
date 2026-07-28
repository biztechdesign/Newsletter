"""Creates content_template.xlsx with formatting."""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Newsletter Content"

# ── Styles ──────────────────────────────────────────────────────────────
TEAL       = "FF34AFBF"
TEAL_LIGHT = "FFE6F7F9"
GRAY_DARK  = "FF2D2D2D"
GRAY_LIGHT = "FFF5F5F5"
YELLOW     = "FFFFF3CD"
BLUE_LIGHT = "FFE8F4FD"
GREEN_LIGHT= "FFEBF7EE"
PINK_LIGHT = "FFFFF0F0"
PURPLE_LIGHT="FFF3EFFF"
WHITE      = "FFFFFFFF"

def hdr_fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def font(bold=False, color="FF000000", size=10): return Font(bold=bold, color=color, size=size, name="Calibri")
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def thin_border():
    s = Side(style="thin", color="FFCCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Column widths ────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 30
ws.column_dimensions["E"].width = 55

# ── Header row ───────────────────────────────────────────────────────────
headers = ["Field", "Value", "Extra", "Extra2", "Drive Link"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = hdr_fill(TEAL)
    cell.font = font(bold=True, color="FFFFFFFF", size=11)
    cell.alignment = center()
    cell.border = thin_border()
ws.row_dimensions[1].height = 22

# ── Data ─────────────────────────────────────────────────────────────────
# (field, value, extra, extra2, drive_link, bg_color, section_label)
rows = [
    # General
    ("month",                "April",       "",  "",  "", TEAL_LIGHT, "GENERAL"),
    ("year",                 "2026",        "",  "",  "", TEAL_LIGHT, None),
    ("ceo_name",             "Maulik Shah", "",  "",  "", TEAL_LIGHT, None),
    ("ceo_title",            "Founder & CEO","", "",  "", TEAL_LIGHT, None),
    ("ceo_message",          "As we move into the next phase of the year, I want to thank each of you for your dedication.", "", "", "", TEAL_LIGHT, None),
    ("ceo_photo",            "",  "",  "",  "https://drive.google.com/file/d/YOUR_FILE_ID/view", TEAL_LIGHT, None),
    ("campaign_banner",      "",  "",  "",  "https://drive.google.com/file/d/YOUR_FILE_ID/view", TEAL_LIGHT, None),

    # Newly Added Customers (folder URL, OR 1 row per customer with Value=name)
    ("new_customer", "", "", "", "https://github.com/biztechdesign/Newsletter/tree/main/images/new_customers", TEAL_LIGHT, "NEWLY ADDED CUSTOMERS\n(folder URL, OR 1 row per customer with Value=name)"),

    # Delivery Insights (repeatable: 1 row per client — logo left, text right)
    ("delivery", "We are pleased to announce the successful go-live of the ACG platform migration from v16 to v19.", "", "", "https://drive.google.com/file/d/YOUR_FILE_ID/view", YELLOW, "DELIVERY INSIGHTS\n(1 row per entry, Drive Link = logo)"),
    ("delivery", "Second client update goes here as its own row, if there is one this month.", "", "", "https://drive.google.com/file/d/YOUR_FILE_ID/view", YELLOW, None),

    # Acknowledging Excellence (repeatable: 1 row per testimonial — logo + text card)
    ("excellence", "I want to mention that we successfully completed the server migration. Sarvil and Milan are gold for me!", "PaceX Client for Milan and Sarvil:", "", "https://drive.google.com/file/d/YOUR_FILE_ID/view", BLUE_LIGHT, "ACKNOWLEDGING EXCELLENCE\n(1 row per entry: Value=testimonial, Extra=client label, Drive Link=logo)"),
    ("excellence", "Second testimonial goes here as its own row, if there is one this month.", "Another Client Name", "", "https://drive.google.com/file/d/YOUR_FILE_ID/view", BLUE_LIGHT, None),

    # HR Insider (wellness banner optional + up to 3 events, each with its own title + photo folder)
    ("hr_wellness_title",   "Wellness & HR Corner", "", "", "", GREEN_LIGHT, "HR INSIDER"),
    ("hr_wellness_banner",  "",  "",  "",  "https://drive.google.com/file/d/YOUR_FILE_ID/view", GREEN_LIGHT, None),
    ("hr_event", "Equality Day Celebration", "", "", "https://github.com/biztechdesign/Newsletter/tree/main/images/events_1", GREEN_LIGHT, None),
    ("hr_event", "Team Outing",              "", "", "https://github.com/biztechdesign/Newsletter/tree/main/images/events_2", GREEN_LIGHT, None),
    ("hr_event", "Diwali Celebration",        "", "", "https://github.com/biztechdesign/Newsletter/tree/main/images/events_3", GREEN_LIGHT, None),
    ("event_photo",         "",  "",  "",  "https://drive.google.com/file/d/YOUR_FILE_ID/view", GREEN_LIGHT, None),

    # Anniversaries — Value = tab name in the Work Anniversary xlsx (e.g. "March-2026"),
    # Drive Link = link to that xlsx file. Leave Drive Link blank to use the default file ID
    # hardcoded in import_content.py's ANNIV_SHEET_ID.
    ("anniversary", "March-2026", "", "", "https://drive.google.com/file/d/YOUR_FILE_ID/view", PINK_LIGHT, "ANNIVERSARIES\n(Value = tab name in the Work Anniversary file, Drive Link = that file)"),

    # Rewards
    ("award", "Appreciation Award", "Heema Patel",       "Business Development Executive",    "https://drive.google.com/file/d/YOUR_FILE_ID/view", PURPLE_LIGHT, "REWARDS\n(1 row per recipient)"),
    ("award", "Spotlight Award",    "Nawabhusen Solanki","Business Development Executive",    "https://drive.google.com/file/d/YOUR_FILE_ID/view", PURPLE_LIGHT, None),
    ("award", "Team Spirit Award",  "Uttam Jain",        "Lead Odoo Consultant",              "https://drive.google.com/file/d/YOUR_FILE_ID/view", PURPLE_LIGHT, None),
    ("award", "Team Spirit Award",  "Pooja Dahale",      "Sr. Odoo Functional Consultant",    "https://drive.google.com/file/d/YOUR_FILE_ID/view", PURPLE_LIGHT, None),

    # Employee Certifications (folder URL, OR 1 row per certificate with Value=title)
    ("employee_certification", "", "", "", "https://github.com/biztechdesign/Newsletter/tree/main/images/certification", BLUE_LIGHT, "CERTIFICATION\n(folder URL, OR 1 row per certificate with Value=title)"),

    # New Additions (new joinees) — recommended: one folder URL, images named "Name--Designation.jpg"
    ("new_joinee", "", "", "", "https://github.com/biztechdesign/Newsletter/tree/main/images/new_joinees", GRAY_LIGHT, "NEW ADDITIONS\n(folder URL, OR 1 row per person with Value=name)"),

    # New Openings
    ("opening", "SEO Executive",               "2", "1+ Years",    "", GRAY_LIGHT, "NEW OPENINGS\n(rows below, OR link an external sheet via new_openings_sheet)"),
    ("opening", "Customer Success Executive",  "1", "1 to 2 Years","", GRAY_LIGHT, None),
    ("opening", "Business Development Intern", "5", "Fresher",     "", GRAY_LIGHT, None),
    ("new_openings_sheet", "", "", "", "https://docs.google.com/spreadsheets/d/YOUR_SHEET_ID/edit", GRAY_LIGHT, "(optional) external New Openings sheet — Drive Link only; overrides the 'opening' rows above when present. Sheet must have columns Month | Position | Experience | Opening, filtered to this month."),

    # Marketing / Blogs
    ("blog", "Custom AI Development vs. Generic Tools",   "https://www.biztechcs.com/blog/custom-ai/",  "", "https://drive.google.com/file/d/YOUR_FILE_ID/view", TEAL_LIGHT, "MARKETING BLOGS"),
    ("blog", "Top 10 Ways AI is Cutting Business Costs",  "https://www.biztechcs.com/blog/ai-costs/",   "", "https://drive.google.com/file/d/YOUR_FILE_ID/view", TEAL_LIGHT, None),

    # Upcoming Event Announcement (folder URL, OR 1 row per announcement with Value=title)
    ("new_announcement", "", "", "", "https://github.com/biztechdesign/Newsletter/tree/main/images/announcement", BLUE_LIGHT, "UPCOMING EVENT ANNOUNCEMENT\n(folder URL, OR 1 row per announcement with Value=title)"),
]

# ── Column header legend (row 2) ─────────────────────────────────────────
legend = [
    ("", "text / title / years / job title / blog title",
     "names (anniversary) / recipient name (award) / count (opening) / blog URL",
     "designation (award) / experience (opening)",
     "Drive link to image"),
]
legend_row = 2
for col, val in enumerate(legend[0], 1):
    cell = ws.cell(row=legend_row, column=col, value=val)
    cell.fill = hdr_fill("FFE0E0E0")
    cell.font = font(bold=False, color="FF666666", size=9)
    cell.alignment = left()
    cell.border = thin_border()
ws.row_dimensions[legend_row].height = 28

# ── Write data rows ───────────────────────────────────────────────────────
START_ROW = 3
prev_section = None
for i, (field, val, extra, extra2, drive, bg, section) in enumerate(rows):
    r = START_ROW + i

    # Section label in column A when it changes
    if section and section != prev_section:
        ws.cell(row=r, column=1, value=field)
        # Add section divider note as a comment-style value in a merged feel
        prev_section = section
    else:
        ws.cell(row=r, column=1, value=field)

    values = [field, val, extra, extra2, drive]
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=r, column=col, value=v)
        cell.fill = hdr_fill(bg)
        cell.font = font(size=10)
        cell.alignment = left()
        cell.border = thin_border()
        if col == 1:  # Field column bold
            cell.font = font(bold=True, size=10)
        if col == 5 and v.startswith("http"):  # Drive link blue
            cell.font = Font(color="FF1155CC", size=10, name="Calibri", underline="single")

    ws.row_dimensions[r].height = 18

# ── Section divider rows ──────────────────────────────────────────────────
# Insert coloured section headers above each group
section_starts = {}
for i, (_, _, _, _, _, _, section) in enumerate(rows):
    if section and section not in section_starts:
        section_starts[section] = START_ROW + i

# Insert rows from bottom to top to preserve indices
for section, row_num in sorted(section_starts.items(), key=lambda x: -x[1]):
    ws.insert_rows(row_num)
    cell = ws.cell(row=row_num, column=1, value=f"▸ {section}")
    cell.fill = hdr_fill("FF2D2D2D")
    cell.font = font(bold=True, color="FFFFFFFF", size=10)
    cell.alignment = left()
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
    ws.row_dimensions[row_num].height = 20

# ── Freeze header rows ────────────────────────────────────────────────────
ws.freeze_panes = "A3"

# ── Save ─────────────────────────────────────────────────────────────────
wb.save(r"C:\Lopa\Newsletter\content_template.xlsx")
print("Saved: content_template.xlsx")
