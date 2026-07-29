"""
create_master_template.py
=========================
Builds the master newsletter sheet that gets shared with every contributor.

It keeps the familiar master layout — Sr No. / Newsletter Sections / POC /
Content — and adds two columns the pipeline reads:

    Photos Folder   where that section's images live under Row Data
    Field           the machine key import_content.py matches on

Contributors only ever touch **Content**. The last two columns are the wiring
and are locked, greyed and can be hidden without breaking anything (hidden
columns are still read).

Usage:
    python create_master_template.py                    -> master_template.xlsx
    python create_master_template.py July-2026/content.xlsx

Then share the file, collect it back, and run:
    python import_content.py --sheet=<file.xlsx>
"""

import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = Path(__file__).parent

# (Sr No., Section name, POC, Field, Row Data folder, what to put in Content)
ROWS = [
    # Which issue this is. Everything keys off these two: the output filename,
    # the Row Data folder that gets read, and the server folder images upload to.
    ("",  "Newsletter Month",                           "",           "month",                  "",                        "e.g. July"),
    ("",  "Newsletter Year",                            "",           "year",                   "",                        "e.g. 2026"),
    (1,  "Message from CXO's Desk",                     "Vedanshi",   "ceo_message",            "Row Data/ceo",            "The CXO's message. Their photo goes in the folder."),
    ("",  "   ↳ CXO Name",                          "Vedanshi",   "ceo_name",               "",                        "Rarely changes, e.g. Maulik Shah"),
    ("",  "   ↳ CXO Title",                         "Vedanshi",   "ceo_title",              "",                        "Rarely changes, e.g. Founder & CEO"),
    (2,  "ISMS Incident Awareness Banner",              "Pratik",     "campaign_banner",        "Row Data/campaign_banner", "Banner image goes in the folder. Leave blank, or NA to drop."),
    (3,  "Monthly Company Expo/Exhibition Photos",      "Parth Pandya", "expo",                 "Row Data/expo",            "Optional heading. Photos go in the folder."),
    (4,  "New Logos Added (New Clients/Projects)",      "Pratik",     "new_customer",           "Row Data/new_customer",    "Logos go in the folder, any filenames."),
    (5,  "Delivery Insights (If any)",                  "Manish & Jay", "delivery",             "Row Data/delivery",        "The write-up. Client logo goes in the folder."),
    (6,  "Response from Client to Employees (If Any)",  "Manish & Jay", "excellence",           "Row Data/excellence",      "The quote. Client logo goes in the folder."),
    (7,  "Customer Testimonials (If received)",         "Manish & Jay", "excellence",           "Row Data/excellence",      "A second quote, same as above."),
    (8,  "Monthly HR Events",                           "Nikita",     "hr_event1",              "Row Data/events/event1",   "Event name. Photos go in the folder."),
    (9,  "Monthly HR Events (2nd event)",               "Nikita",     "hr_event2",              "Row Data/events/event2",   "Second event, if there is one."),
    (10, "Monthly HR Events (3rd event)",               "Nikita",     "hr_event3",              "Row Data/events/event3",   "Third event, if there is one."),
    (11, "Wellness & HR Corner",                        "HR",         "hr_wellness_banner",     "Row Data/hr_wellness",     "Heading. Banner image goes in the folder."),
    (12, "R & R",                                       "Nikita",     "award",                  "Row Data/awards",          "Names come from the filenames -- see the note below."),
    (13, "New Joiners (This Month)",                    "Nikita",     "new_joinee",             "Row Data/new_joinee",      "Names come from the filenames -- see the note below."),
    (14, "Work Anniversary (This Month)",               "Nikita",     "anniversary",            "",                         "Paste from your sheet: Name <tab> 5 years, one per line."),
    (15, "Employee Certifications (This Month)",        "Nikita",     "employee_certification", "Row Data/certification",   "Certificate images go in the folder."),
    (16, "New Openings (This Month)",                   "Nikita",     "new_openings_sheet",     "",                         "Paste with header: Position <tab> Experience <tab> Opening."),
    (17, "Employee Training (This Month)",              "Nikita",     "employee_training",      "Row Data/training",        "Optional heading. Photos go in the folder."),
    (18, "Employee Workshop (This Month)",              "Nikita",     "employee_workshop",      "Row Data/workshop",        "Optional heading. Photos go in the folder."),
    (19, "Announcement of any Events",                  "Nikita",     "new_announcement",       "Row Data/announcement",    "Event name. Banner goes in the folder."),
    (20, "Marketing Highlights (Blogs, Case Studies…)", "Dhaval",     "marketing_highlights",   "Row Data/marketing/blogs", "One per line: Title <tab> https://link"),
]

HEADERS = ["Sr No.", "Newsletter Sections", "POC", "Content", "Photos Folder", "Field", "Note"]
WIDTHS = [8, 42, 16, 62, 26, 24, 40]

# Columns the importer reads; everything else is for people. Note is free text
# and is never parsed, so anything can be written there without side effects.
WIRING_COLUMNS = (5, 6)

BLACK = PatternFill("solid", fgColor="000000")
GREY = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build(out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Newsletter Content"

    # ── Instructions above the table ──────────────────────────────────────
    # Everything essential is stated here, in plain cells, because this block
    # survives being copy-pasted into another sheet while cell comments,
    # colours and column widths do not.
    notes = [
        "BIZTECH NEWSLETTER — fill in the Content column only. Leave the grey Photos Folder / Field columns alone.",
        "Nothing this month? Leave Content blank or write NA — that section is then left out of the newsletter entirely.",
        "PHOTOS: drop them straight into the folder named in 'Photos Folder'. No links needed. Order follows the filename, so prefix 01, 02, … where it matters.",
        "        R & R filename:       01 - Award Title - Person Name - Designation.jpg   (repeat the award title to group several people under one award)",
        "        New Joiners filename: Person Name - Designation.jpg          iPhone HEIC photos must be re-saved as PNG or JPEG first — renaming them is not enough.",
        "PASTE-IN LISTS (one per line, Tab between the parts — paste straight from another sheet):",
        "        Work Anniversary:  Person Name <Tab> 5 years",
        "        New Openings:      Position <Tab> Experience <Tab> Opening        (keep the header line)",
        "        Marketing:         Blog Title <Tab> https://link",
    ]
    for i, note in enumerate(notes, start=1):
        cell = ws.cell(row=i, column=1, value=note)
        cell.font = Font(bold=(i == 1), size=11 if i == 1 else 9,
                         color="000000" if i == 1 else "555555")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(HEADERS))

    header_row = len(notes) + 2

    # ── Header ────────────────────────────────────────────────────────────
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.fill = BLACK
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = WIDTHS[col - 1]

    # ── Rows ──────────────────────────────────────────────────────────────
    for n, (sr, section, poc, field, folder, hint) in enumerate(ROWS):
        r = header_row + 1 + n
        values = [sr, section, poc, None, folder, field, None]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                horizontal="center" if col in (1, 3) else "left",
                wrap_text=col in (2, 4, 7),
            )
            # The two wiring columns are not for contributors to edit.
            if col in WIRING_COLUMNS:
                cell.fill = GREY
                cell.font = Font(size=9, color="777777")
        # Guidance hangs off the Content cell as a hover note rather than a
        # column of its own — it stays available without adding a column
        # people have to read past, and it leaves the cell genuinely empty
        # (empty is what "no content this month" means to the importer).
        note = Comment(hint, "Newsletter")
        note.width, note.height = 320, 60
        ws.cell(row=r, column=4).comment = note
        ws.row_dimensions[r].height = 30

    # Nudge people toward NA rather than inventing their own wording.
    dv = DataValidation(type="list", formula1='"NA"', allow_blank=True, showDropDown=False)
    dv.prompt = "Type the content, leave blank, or write NA to drop this section."
    dv.promptTitle = "Content"
    ws.add_data_validation(dv)
    dv.add(f"D{header_row + 1}:D{header_row + len(ROWS)}")

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path, header_row


MONTH_FOLDER = re.compile(
    r"^(January|February|March|April|May|June|July|August|September|October|"
    r"November|December)-\d{4}$", re.IGNORECASE
)


def ensure_folders(sheet_path: Path) -> list[str]:
    """
    Create the Row Data folder for every section the sheet names, next to the
    sheet itself. Nested paths (events/event1, marketing/blogs) are created in
    full. Existing folders and their contents are left alone.

    Only done when the sheet actually lives in a "<Month>-<Year>" folder. The
    blank template is written to the project root, and scaffolding there would
    litter it with twenty empty directories that belong to no issue.
    """
    if not MONTH_FOLDER.match(sheet_path.parent.name):
        return []
    root = sheet_path.parent / "Row Data"
    made = []
    for _sr, _section, _poc, _field, folder, _hint in ROWS:
        if not folder:
            continue
        target = root / folder.split("/", 1)[1] if folder.startswith("Row Data/") else root / folder
        if not target.is_dir():
            target.mkdir(parents=True, exist_ok=True)
            made.append(target.relative_to(root).as_posix())
    return made


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else BASE_DIR / "master_template.xlsx"
    if not target.is_absolute():
        target = BASE_DIR / target
    path, header_row = build(target)
    print(f"Master template written: {path}")
    print(f"  {len(ROWS)} sections, header on row {header_row}")
    print(f"  Contributors fill column D (Content); columns E/F are the wiring.")
    made = ensure_folders(path)
    if made:
        print(f"  Created {len(made)} photo folder(s): {', '.join(made)}")
    print(f"\nNext: share it, collect it back, then run")
    print(f"  python import_content.py --sheet={target.name}")
